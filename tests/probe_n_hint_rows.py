from urllib.request import Request, urlopen
from openpyxl import load_workbook
import re

WORKBOOK = "部分.工作簿1.xlsx"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://openstd.samr.gov.cn/",
}


def fetch(url: str) -> tuple[str, str]:
    req = Request(url, headers=HEADERS)
    with urlopen(req, timeout=25) as resp:
        data = resp.read().decode("utf-8", errors="replace")
        return data, resp.geturl()


def extract_hcno_from_new_gb_info(html: str) -> str | None:
    m = re.search(r'data-value="([A-F0-9]{32})"\s+class="btn fk_btn', html)
    if m:
        return m.group(1)
    return None


def marks(html: str) -> dict[str, bool]:
    text = html
    return {
        "has_采标情况": "采标情况" in text,
        "has_IEC": "IEC" in text,
        "has_等同采用": "等同采用" in text,
        "has_login_word": ("登录" in text) or ("login" in text.lower()) or ("uac.sacinfo.org.cn" in text),
    }


def get_n_hint_urls(limit: int = 6) -> list[tuple[int, str]]:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    out: list[tuple[int, str]] = []
    for r in range(2, ws.max_row + 1):
        n_val = ws[f"N{r}"].value
        o_val = ws[f"O{r}"].value
        if n_val and "采" in str(n_val) and o_val:
            out.append((r, str(o_val).strip()))
            if len(out) >= limit:
                break
    return out


def main() -> None:
    rows = get_n_hint_urls(limit=6)
    print(f"n_hint_samples={len(rows)}")
    for row, new_url in rows:
        try:
            html_new, final_new = fetch(new_url)
            hcno = extract_hcno_from_new_gb_info(html_new)
            review_url = f"https://openstd.samr.gov.cn/bzgk/gb/review?hcno={hcno}" if hcno else ""
            if review_url:
                html_review, final_review = fetch(review_url)
                m = marks(html_review)
            else:
                final_review = ""
                m = {}

            print("\n---")
            print(f"row={row}")
            print(f"new_final={final_new}")
            print(f"review_url={review_url}")
            print(f"review_final={final_review}")
            print(f"marks={m}")
        except Exception as exc:
            print("\n---")
            print(f"row={row}")
            print(f"error={exc}")


if __name__ == "__main__":
    main()
