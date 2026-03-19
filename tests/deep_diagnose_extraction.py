"""
深度诊断：为什么即使获取了review页面，也提取不到采标信息？

可能的原因：
1. extract_caibiao_info 正则表达式不匹配
2. review页面的HTML结构不同
3. 采标信息在隐藏的div中
"""

import sys
import re
from pathlib import Path
from urllib.request import urlopen, Request

sys.path.insert(0, str(Path(__file__).parent.parent))

import fetch
from openpyxl import load_workbook


def diagnose_extraction_failure():
    """诊断为什么提取失败"""
    print("=" * 70)
    print("深度诊断：采标信息提取失败原因")
    print("=" * 70)
    
    # 获取一个实际的review页面
    try:
        excel_path = Path(__file__).parent.parent / "部分.工作簿1.xlsx"
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 取第一个有URL的行
        for row in range(2, min(10, ws.max_row + 1)):
            url = ws[f"O{row}"].value
            if url and isinstance(url, str) and url.startswith("http"):
                test_url = url
                test_row = row
                break
        
        print(f"\n测试URL（行{test_row}）：{test_url[:60]}...")
        
        # 获取newGbInfo页面
        print("\n【获取newGbInfo页面】")
        headers = {
            'User-Agent': 'Mozilla/5.0',
            'Referer': 'https://openstd.samr.gov.cn/'
        }
        req = Request(test_url, headers=headers)
        resp = urlopen(req, timeout=5)
        newgb_content = resp.read().decode('utf-8', errors='replace')
        
        print(f"  页面大小：{len(newgb_content)} bytes")
        
        # 检查是否真的包含内容
        if not newgb_content or len(newgb_content) < 100:
            print(f"  ⚠️  页面内容过少或为空！")
            print(f"  内容预览：{newgb_content[:200]}")
            return False
        
        # 检查是否包含"采标"
        if "采标" in newgb_content:
            print("  ✓ 页面包含'采标'关键词")
        else:
            print("  ✗ 页面不包含'采标'关键词")
        
        # 检查页面结构
        print("\n【分析newGbInfo页面结构】")
        print(f"  包含<h2: {newgb_content.count('<h2')}")
        print(f"  包含class=\"title-text\": {newgb_content.count('class=\"title-text\"')}")
        print(f"  包含'采标情况': {newgb_content.count('采标情况')}")
        
        # 显示所有<h2>标题
        h2_pattern = r'<h2[^>]*class="title-text"[^>]*>(.*?)</h2>'
        h2_matches = re.findall(h2_pattern, newgb_content, flags=re.IGNORECASE)
        print(f"\n  找到 {len(h2_matches)} 个section标题：")
        for i, title in enumerate(h2_matches[:5], 1):
            clean_title = fetch.html_to_text(f"<h2>{title}</h2>")
            print(f"    {i}. {clean_title}")
        
        # 尝试extract
        print("\n【尝试从newGbInfo提取】")
        has_module, caibiao_text = fetch.extract_caibiao_info(newgb_content)
        print(f"  has_module: {has_module}")
        print(f"  caibiao_text长度: {len(caibiao_text)}")
        if caibiao_text:
            print(f"  内容预览: {caibiao_text[:100]}...")
        
        if not caibiao_text:
            print(f"  ✗ newGbInfo无采标信息，需要获取review")
            
            # 获取review URL
            hcnos = fetch.extract_hcnos_for_review(test_url, newgb_content)
            if hcnos:
                first_hcno = hcnos[0]
                review_url = fetch.build_review_url(test_url, first_hcno)
                
                print(f"\n【获取review页面】")
                print(f"  review URL: {review_url[:60]}...")
                
                req = Request(review_url, headers=headers)
                resp = urlopen(req, timeout=5)
                review_content = resp.read().decode('utf-8', errors='replace')
                
                print(f"  页面大小：{len(review_content)} bytes")
                print(f"  包含'采标': {review_content.count('采标')}")
                print(f"  包含'采用': {review_content.count('采用')}")
                print(f"  包含'标准': {review_content.count('标准')}")
                
                # 检查采标情况section
                section_text = fetch.extract_section_text_by_title(review_content, "采标情况")
                print(f"\n【尝试extract_section_text_by_title】")
                print(f"  提取到的section文本长度: {len(section_text)}")
                if section_text:
                    print(f"  内容预览: {section_text[:150]}...")
                else:
                    print(f"  ✗ 无法找到'采标情况'section")
                    
                    # 尝试手动查找<h2>采标情况</h2>
                    print(f"\n【手动查找'采标情况'】")
                    if '<h2' in review_content and '采标情况' in review_content:
                        # 找到位置
                        idx = review_content.find('采标情况')
                        context = review_content[max(0, idx-200):idx+300]
                        print(f"  上下文：...{context}...")
                    else:
                        print(f"  ✗ 页面不包含'采标情况'")
                
                # 尝试从review提取
                print(f"\n【尝试从review提取】")
                r_has_module, r_caibiao_text = fetch.extract_caibiao_info(review_content)
                print(f"  has_module: {r_has_module}")
                print(f"  caibiao_text长度: {len(r_caibiao_text)}")
                if r_caibiao_text:
                    print(f"  内容预览: {r_caibiao_text[:100]}...")
                else:
                    print(f"  ✗ review页面也无法提取采标")
                    
                    # 尝试改进的extract
                    print(f"\n【尝试改进的extraction方法】")
                    
                    # 直接查找采用/等同采用开头的句子
                    patterns = [
                        r"(本(?:标准|文件)[^。；]{0,160}(?:等同采用|修改采用|非等效采用|采用)[^。；]{0,240}[。；]?)",
                        r"(采用(?:IEC|ISO|EN|ASTM)[^。；]{0,240}[。；]?)",
                        r"(<h2[^>]*>\s*采标情况\s*</h2>)",
                    ]
                    
                    for i, pattern in enumerate(patterns, 1):
                        matches = re.findall(pattern, review_content, flags=re.IGNORECASE)
                        print(f"  模式{i}：找到 {len(matches)} 个匹配")
                        if matches:
                            for match in matches[:1]:
                                preview = match[:80] if len(match) < 80 else match[:80] + "..."
                                print(f"    {preview}")
            else:
                print(f"  ✗ 无法找到review URL")
        
        return True
        
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    diagnose_extraction_failure()
