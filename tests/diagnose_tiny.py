"""
超快速诊断：只测试5个URL，快速判断方向
"""

import sys
import time
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook


def get_urls(excel_path, count=5):
    """获取前几个URL"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    urls = []
    
    for row in range(2, ws.max_row + 1):
        url = ws['O' + str(row)].value
        if url and isinstance(url, str) and url.startswith('http'):
            urls.append(url)
            if len(urls) >= count:
                break
    
    return urls


def fetch_url(url, timeout=6):
    """发起单个请求"""
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        req = Request(url, headers=headers)
        resp = urlopen(req, timeout=timeout)
        return {'status': resp.status, 'error': None}
    except HTTPError as e:
        return {'status': e.code, 'error': f'HTTP{e.code}'}
    except Exception as e:
        return {'status': None, 'error': type(e).__name__}


if __name__ == '__main__':
    excel = sys.argv[1] if len(sys.argv) > 1 else '部分.工作簿1.xlsx'
    
    print("=" * 60)
    print("超快速诊断（只测5个URL）")
    print("=" * 60)
    
    urls = get_urls(excel, count=5)
    print(f"\n加载了 {len(urls)} 个URL\n")
    
    # 测试1：单线程
    print("【测试1】单线程无延迟...")
    single_results = []
    for i, url in enumerate(urls, 1):
        result = fetch_url(url)
        single_results.append(result)
        status = result['status'] if result['status'] else result['error']
        print(f"  {i}. {status}")
    
    time.sleep(1)
    
    # 测试2：多线程
    print("\n【测试2】多线程(10个工作线程)...")
    multi_results = [None] * len(urls)
    
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_url, url): i for i, url in enumerate(urls)}
        for future in as_completed(futures):
            i = futures[future]
            multi_results[i] = future.result()
    
    for i, result in enumerate(multi_results, 1):
        status = result['status'] if result['status'] else result['error']
        print(f"  {i}. {status}")
    
    # 分析
    print("\n" + "=" * 60)
    print("分析")
    print("=" * 60)
    
    single_401 = sum(1 for r in single_results if r['status'] == 401)
    multi_401 = sum(1 for r in multi_results if r['status'] == 401)
    single_200 = sum(1 for r in single_results if r['status'] == 200)
    multi_200 = sum(1 for r in multi_results if r['status'] == 200)
    
    single_401_rate = (single_401 / len(urls) * 100) if urls else 0
    multi_401_rate = (multi_401 / len(urls) * 100) if urls else 0
    diff = multi_401_rate - single_401_rate
    
    print(f"\n单线程：200={single_200}个, 401={single_401}个 (401率: {single_401_rate:.0f}%)")
    print(f"多线程：200={multi_200}个, 401={multi_401}个 (401率: {multi_401_rate:.0f}%)")
    print(f"\n401率差异：{diff:+.0f}%")
    
    print("\n【结论】：")
    if single_401_rate >= 80:
        print("⚠️  极可能是 IP 被封禁")
        print("   → 需要更换IP或等待")
    elif diff > 20:
        print("⚠️  主要是 并发限制")
        print("   → 改workers=3, 加延迟0.5s")
    elif single_401_rate > 50:
        print("⚠️  可能是IP/速率混合问题")
        print("   → 先测试延迟是否有帮助")
    else:
        print("✓ 暂时没问题，可以运行full run")
    
    print("\n" + "=" * 60)
