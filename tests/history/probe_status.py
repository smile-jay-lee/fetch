from collections import Counter
import re
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

wb = load_workbook("部分.工作簿1.xlsx", data_only=True)
ws = wb.active
rows: list[tuple[int, str]] = []
for r in range(2, ws.max_row + 1):
    n = ws[f"N{r}"].value
    o = ws[f"O{r}"].value
    if n and "采" in str(n) and o:
        m = re.search(r"hcno=([A-F0-9]{32})", str(o))
        if m:
            rows.append((r, m.group(1)))

headers = {
    "User-Agent": "Mozilla/5.0",
    "Referer": "https://openstd.samr.gov.cn/",
    "Accept-Language": "zh-CN,zh;q=0.9",
}

status = Counter()
accessible_with_caibiao = 0
for row, hcno in rows:
    u = f"https://openstd.samr.gov.cn/bzgk/gb/review?hcno={hcno}"
    try:
        resp = urlopen(Request(u, headers=headers), timeout=20)
        html = resp.read().decode("utf-8", errors="replace")
        status[str(resp.status)] += 1
        if "采标情况" in html:
            accessible_with_caibiao += 1
    except HTTPError as e:
        status[f"HTTP_{e.code}"] += 1
    except URLError:
        status["URLError"] += 1

print("n_hint_rows", len(rows))
print("status_counts", dict(status))
print("accessible_with_caibiao", accessible_with_caibiao)
