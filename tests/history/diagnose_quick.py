"""
快速诊断脚本：快速判断是IP封禁还是并发限制

只运行2个核心测试：
1. 单线程无延迟 - 基准
2. 多线程(10工作线程)无延迟 - 当前默认配置

这样可以快速判断：
- 如果401率都很高 → IP可能被封
- 如果10线程401率 > 单线程401率 > 30% → 并发限制
"""

import sys
import time
import json
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime
import statistics

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


class QuickDiagnoser:
    """快速诊断HTTP请求阻塞原因"""

    def __init__(self, workbook_path: str, sample_size: int = 20):
        self.workbook_path = workbook_path
        self.sample_size = sample_size
        self.urls = self._load_urls_from_excel()
        self.timeout = 8  # 缩短超时时间

    def _load_urls_from_excel(self) -> list:
        """从Excel的O列加载URL"""
        try:
            wb = load_workbook(self.workbook_path, data_only=True)
            ws = wb.active
            urls = []
            
            for row in range(2, ws.max_row + 1):
                url = ws['O' + str(row)].value
                if url and isinstance(url, str) and url.startswith('http'):
                    urls.append((row, url))
                    
            if len(urls) > self.sample_size:
                step = len(urls) // self.sample_size
                urls = urls[::step][:self.sample_size]
            
            print(f"✓ 加载了 {len(urls)} 个URL")
            return urls
        except Exception as e:
            print(f"✗ 加载失败: {e}")
            sys.exit(1)

    def _fetch_single(self, url: str) -> dict:
        """发起单个请求"""
        start_time = time.time()
        result = {'url': url, 'status': None, 'error': None, 'response_time': 0}
        
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            req = Request(url, headers=headers)
            response = urlopen(req, timeout=self.timeout)
            result['status'] = response.status
        except HTTPError as e:
            result['status'] = e.code
            result['error'] = f"HTTP {e.code}"
        except (URLError, TimeoutError):
            result['error'] = 'Timeout'
        except KeyboardInterrupt:
            raise
        except Exception as e:
            result['error'] = type(e).__name__
        
        result['response_time'] = time.time() - start_time
        return result

    def test_single_thread(self) -> dict:
        """单线程测试"""
        print("\n[测试1/2] 单线程无延迟...")
        start = time.time()
        results = []
        
        for i, (row_num, url) in enumerate(self.urls):
            result = self._fetch_single(url)
            results.append(result)
            if (i + 1) % 5 == 0:
                print(f"  {i+1}/{len(self.urls)}", end='\r')
        
        print(f"  {len(self.urls)}/{len(self.urls)}  ✓")
        return {'name': '单线程', 'workers': 1, 'elapsed': time.time() - start, 'results': results}

    def test_multi_thread(self) -> dict:
        """多线程测试"""
        print("[测试2/2] 多线程(10工作线程)...")
        start = time.time()
        results = [None] * len(self.urls)
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(self._fetch_single, url): (i, url) 
                      for i, (row_num, url) in enumerate(self.urls)}
            
            completed = 0
            for future in as_completed(futures):
                idx, url = futures[future]
                try:
                    results[idx] = future.result()
                except Exception as e:
                    results[idx] = {'url': url, 'status': None, 'error': str(e), 'response_time': 0}
                
                completed += 1
                if completed % 5 == 0:
                    print(f"  {completed}/{len(self.urls)}", end='\r')
        
        print(f"  {len(self.urls)}/{len(self.urls)}  ✓")
        return {'name': '多线程(10)', 'workers': 10, 'elapsed': time.time() - start, 'results': results}

    def analyze(self, test1: dict, test2: dict):
        """分析结果"""
        def calc_stats(results):
            status_401 = sum(1 for r in results if r and r['status'] == 401)
            status_200 = sum(1 for r in results if r and r['status'] == 200)
            errors = sum(1 for r in results if r and r['error'])
            total = len([r for r in results if r])
            rate_401 = (status_401 / total * 100) if total > 0 else 0
            return {
                '200': status_200,
                '401': status_401,
                '其他错误': errors,
                '总': total,
                '401率': rate_401
            }
        
        stats1 = calc_stats(test1['results'])
        stats2 = calc_stats(test2['results'])
        
        return stats1, stats2

    def print_results(self, test1: dict, test2: dict, stats1: dict, stats2: dict):
        """打印结果"""
        print(f"\n{'='*70}")
        print("快速诊断结果")
        print(f"{'='*70}\n")
        
        print(f"{'测试':<20} {'200':<8} {'401':<8} {'401率':<10} {'耗时':<8}")
        print("-" * 70)
        print(f"{test1['name']:<20} {stats1['200']:<8} {stats1['401']:<8} "
              f"{stats1['401率']:.1f}%{'':<5} {test1['elapsed']:.1f}s")
        print(f"{test2['name']:<20} {stats2['200']:<8} {stats2['401']:<8} "
              f"{stats2['401率']:.1f}%{'':<5} {test2['elapsed']:.1f}s")
        
        print(f"\n401率差异: {stats2['401率'] - stats1['401率']:.1f}%")

    def diagnose(self, stats1: dict, stats2: dict):
        """给出诊断"""
        rate1 = stats1['401率']
        rate2 = stats2['401率']
        diff = rate2 - rate1
        
        print(f"\n{'='*70}")
        print("诊断结论")
        print(f"{'='*70}\n")
        
        if rate1 > 60:
            print("❌ 【极可能】IP被封禁或加入黑名单")
            print(f"   理由：单线程401率{rate1:.1f}% > 60%，即使单个请求也被拒")
            print("\n【建议】：")
            print("   1. 尝试更换IP（VPN/代理）")
            print("   2. 降低请求频率（等1-2小时再试）")
            print("   3. 改变请求特征（User-Agent等）")
        
        elif diff > 20:
            print("✓ 【主要原因】并发/速率限制")
            print(f"   理由：并发从1→10，401率从{rate1:.1f}% → {rate2:.1f}%，上升{diff:.1f}%")
            print("\n【改进方案】：")
            print("   1. 降低workers数（改为2-3个）")
            print("   2. 加请求间延迟（500ms-1s）")
            print("   3. 实现指数退避重试")
            print("\n【快速修改】：")
            print("   在fetch.py中改 --workers 默认值从10改为3")
        
        else:
            print("⚠️  【不明确】需要更详细的诊断")
            print(f"   单线程: {rate1:.1f}% | 10线程: {rate2:.1f}% | 差异: {diff:.1f}%")
            print("\n【建议】：运行完整诊断脚本 diagnose_blocking.py")

    def save_json(self, test1: dict, test2: dict):
        """保存JSON结果"""
        data = {
            'timestamp': datetime.now().isoformat(),
            'tests': [test1, test2]
        }
        output = 'diagnose_quick_results.json'
        with open(output, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        print(f"\n✓ 详细结果保存到: {output}")


if __name__ == '__main__':
    excel_file = sys.argv[1] if len(sys.argv) > 1 else '部分.工作簿1.xlsx'
    sample_size = int(sys.argv[2]) if len(sys.argv) > 2 else 20
    
    if not Path(excel_file).exists():
        print(f"✗ 文件不存在: {excel_file}")
        sys.exit(1)
    
    print(f"{'='*70}")
    print("快速诊断：IP封禁 vs 并发限制")
    print(f"{'='*70}\n")
    
    diagnoser = QuickDiagnoser(excel_file, sample_size)
    
    try:
        test1 = diagnoser.test_single_thread()
        time.sleep(1)
        test2 = diagnoser.test_multi_thread()
        
        stats1, stats2 = diagnoser.analyze(test1, test2)
        diagnoser.print_results(test1, test2, stats1, stats2)
        diagnoser.diagnose(stats1, stats2)
        diagnoser.save_json(test1, test2)
        
        print(f"\n{'='*70}\n")
    
    except KeyboardInterrupt:
        print("\n\n✗ 用户中止诊断")
    except Exception as e:
        print(f"\n✗ 错误: {e}")
        import traceback
        traceback.print_exc()
