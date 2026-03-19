"""
关键发现：页面大小为0 bytes
可能原因：
1. HTTP response为empty body
2. response body被读取两次导致缓冲区为空
3. 网络问题
"""

import sys
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import HTTPError

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook


print("=" * 70)
print("诊断：HTTP响应为什么是空的？")
print("=" * 70)

try:
    excel_path = Path(__file__).parent.parent / "部分.工作簿1.xlsx"
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 获取第一个有效URL
    test_url = None
    for row in range(2, min(10, ws.max_row + 1)):
        url = ws[f"O{row}"].value
        if url and isinstance(url, str) and url.startswith("http"):
            test_url = url
            break
    
    if not test_url:
        print("❌ 没找到有效URL")
        sys.exit(1)
    
    print(f"\nURL：{test_url}")
    
    # 设置headers
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://openstd.samr.gov.cn/'
    }
    
    # 发起请求
    print("\n【发起HTTP请求】")
    req = Request(test_url, headers=headers)
    
    try:
        resp = urlopen(req, timeout=10)
        
        # 检查响应
        print(f"  HTTP状态码：{resp.status}")
        print(f"  Content-Type：{resp.headers.get('Content-Type', 'unknown')}")
        print(f"  Content-Length：{resp.headers.get('Content-Length', 'unknown')}")
        print(f"  Content-Encoding：{resp.headers.get('Content-Encoding', 'none')}")
        
        # 读取body
        print("\n【读取响应体】")
        body = resp.read()
        print(f"  读取到的字节数：{len(body)}")
        
        if len(body) == 0:
            print("  ⚠️  响应体为空！")
            print(f"\n  响应头信息：")
            for key, value in resp.headers.items():
                print(f"    {key}: {value}")
        else:
            # 尝试解码
            try:
                content = body.decode('utf-8', errors='replace')
                print(f"  解码后字符数：{len(content)}")
                
                # 检查内容
                if '<!DOCTYPE' in content or '<html' in content:
                    print(f"  ✓ 有效的HTML内容")
                    print(f"  内容预览：{content[:200]}...")
                else:
                    print(f"  内容预览：{content[:200]}...")
            except Exception as e:
                print(f"  ❌ 解码失败：{e}")
    
    except HTTPError as http_err:
        print(f"  ❌ HTTP错误：{http_err.code} {http_err.reason}")
        print(f"      URL：{http_err.url}")
        print(f"      Headers：{dict(http_err.headers)}")
    
except Exception as e:
    print(f"❌ 错误：{e}")
    import traceback
    traceback.print_exc()
