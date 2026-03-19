"""
诊断脚本：分析是IP封禁还是并发高导致的HTTP 401问题

测试场景：
1. 单线程无延迟 - 快速顺序请求
2. 单线程有延迟 - 每个请求间隔500ms
3. 单线程大延迟 - 每个请求间隔1s
4. 多线程(3个工作线程) - 低并发
5. 多线程(10个工作线程) - 当前默认并发
6. 多线程(20个工作线程) - 高并发

分析指标：
- HTTP 401/200/其他状态分布
- 超时和连接错误
- 如果401率随并发数增加而明显增加 → 并发限制
- 如果单线程也有高401率 → IP封禁
- 如果延迟能降低401率 → 目标网站有速率限制
"""

import sys
import time
import json
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime
import statistics

# 添加fetch模块路径
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


class RequestDiagnoser:
    """诊断HTTP请求阻塞原因"""

    def __init__(self, workbook_path: str, sample_size: int = 30):
        """
        初始化诊断器
        
        Args:
            workbook_path: Excel文件路径
            sample_size: 采样多少行进行测试（减少测试时间）
        """
        self.workbook_path = workbook_path
        self.sample_size = sample_size
        self.urls = self._load_urls_from_excel()
        self.results = {}
        self.timeout = 10

    def _load_urls_from_excel(self) -> list:
        """从Excel的O列加载URL"""
        try:
            wb = load_workbook(self.workbook_path, data_only=True)
            ws = wb.active
            urls = []
            
            for row in range(2, ws.max_row + 1):
                url = ws['O' + str(row)].value
                if url and isinstance(url, str) and url.startswith('http'):
                    urls.append((row, url))
                    
            # 采样
            if len(urls) > self.sample_size:
                step = len(urls) // self.sample_size
                urls = urls[::step][:self.sample_size]
            
            print(f"✓ 从Excel加载了 {len(urls)} 个URL（总共 {ws.max_row - 1} 行）")
            return urls
        except Exception as e:
            print(f"✗ 加载Excel文件失败: {e}")
            sys.exit(1)

    def _fetch_single(self, url: str, delay_before: float = 0) -> dict:
        """
        发起单个请求
        
        Args:
            url: 请求URL
            delay_before: 请求前延迟（秒）
            
        Returns:
            结果字典: {status, error, response_time}
        """
        if delay_before > 0:
            time.sleep(delay_before)
        
        start_time = time.time()
        result = {
            'url': url,
            'status': None,
            'error': None,
            'response_time': 0
        }
        
        try:
            # 发起请求
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            req = Request(url, headers=headers)
            response = urlopen(req, timeout=self.timeout)
            result['status'] = response.status
            result['response_time'] = time.time() - start_time
            
        except HTTPError as e:
            result['status'] = e.code
            result['error'] = f"HTTP {e.code}"
            result['response_time'] = time.time() - start_time
        except (URLError, TimeoutError) as e:
            result['error'] = 'Timeout' if isinstance(e, TimeoutError) else 'URLError'
            result['response_time'] = time.time() - start_time
        except KeyboardInterrupt:
            raise
        except Exception as e:
            result['error'] = str(type(e).__name__)
            result['response_time'] = time.time() - start_time
        
        return result

    def test_single_thread_no_delay(self) -> dict:
        """单线程无延迟测试"""
        print("\n[1/6] 测试场景: 单线程无延迟（快速顺序）...")
        start = time.time()
        results = []
        
        for row_num, url in self.urls:
            result = self._fetch_single(url, delay_before=0)
            results.append(result)
            # 打印进度
            if len(results) % 5 == 0:
                print(f"  进度: {len(results)}/{len(self.urls)}")
        
        elapsed = time.time() - start
        return {
            'name': '单线程无延迟',
            'config': {'workers': 1, 'delay_ms': 0},
            'elapsed': elapsed,
            'results': results
        }

    def test_single_thread_500ms_delay(self) -> dict:
        """单线程500ms延迟测试"""
        print("\n[2/6] 测试场景: 单线程500ms延迟...")
        start = time.time()
        results = []
        
        for row_num, url in self.urls:
            result = self._fetch_single(url, delay_before=0.5)
            results.append(result)
            if len(results) % 5 == 0:
                print(f"  进度: {len(results)}/{len(self.urls)}")
        
        elapsed = time.time() - start
        return {
            'name': '单线程500ms延迟',
            'config': {'workers': 1, 'delay_ms': 500},
            'elapsed': elapsed,
            'results': results
        }

    def test_single_thread_1s_delay(self) -> dict:
        """单线程1s延迟测试"""
        print("\n[3/6] 测试场景: 单线程1s延迟...")
        start = time.time()
        results = []
        
        for row_num, url in self.urls:
            result = self._fetch_single(url, delay_before=1.0)
            results.append(result)
            if len(results) % 5 == 0:
                print(f"  进度: {len(results)}/{len(self.urls)}")
        
        elapsed = time.time() - start
        return {
            'name': '单线程1s延迟',
            'config': {'workers': 1, 'delay_ms': 1000},
            'elapsed': elapsed,
            'results': results
        }

    def test_multi_thread(self, workers: int, delay_ms: int = 0) -> dict:
        """多线程测试"""
        delay_s = delay_ms / 1000
        print(f"\n[多线程] 测试场景: {workers}个工作线程, {delay_ms}ms延迟...")
        start = time.time()
        results = []
        
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(self._fetch_single, url, delay_s): url
                for row_num, url in self.urls
            }
            
            completed = 0
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                completed += 1
                if completed % 5 == 0:
                    print(f"  进度: {completed}/{len(self.urls)}")
        
        elapsed = time.time() - start
        return {
            'name': f'多线程{workers}工作线程{delay_ms}ms延迟',
            'config': {'workers': workers, 'delay_ms': delay_ms},
            'elapsed': elapsed,
            'results': results
        }

    def run_all_tests(self) -> list:
        """运行所有测试场景"""
        print(f"{'='*60}")
        print(f"诊断脚本开始: {len(self.urls)} 个采样URL")
        print(f"{'='*60}")
        
        all_results = []
        
        # 单线程系列
        all_results.append(self.test_single_thread_no_delay())
        time.sleep(2)  # 测试间隔，让服务器休息
        
        all_results.append(self.test_single_thread_500ms_delay())
        time.sleep(2)
        
        all_results.append(self.test_single_thread_1s_delay())
        time.sleep(2)
        
        # 多线程系列
        all_results.append(self.test_multi_thread(workers=3, delay_ms=0))
        time.sleep(2)
        
        all_results.append(self.test_multi_thread(workers=10, delay_ms=0))
        time.sleep(2)
        
        all_results.append(self.test_multi_thread(workers=20, delay_ms=0))
        
        return all_results

    @staticmethod
    def analyze_results(all_results: list) -> pd.DataFrame:
        """分析结果，生成对比表格"""
        analysis = []
        
        for test_result in all_results:
            results = test_result['results']
            
            # 统计各种状态码
            status_counts = {}
            error_counts = {}
            response_times = []
            
            for r in results:
                if r['status']:
                    status_counts[r['status']] = status_counts.get(r['status'], 0) + 1
                if r['error']:
                    error_counts[r['error']] = error_counts.get(r['error'], 0) + 1
                if r['response_time'] > 0:
                    response_times.append(r['response_time'])
            
            # 计算401率
            http_401_count = status_counts.get(401, 0)
            http_200_count = status_counts.get(200, 0)
            total = len(results)
            http_401_rate = (http_401_count / total * 100) if total > 0 else 0
            
            # 计算响应时间统计
            avg_response_time = statistics.mean(response_times) if response_times else 0
            
            analysis.append({
                '测试场景': test_result['name'],
                '配置': f"workers={test_result['config']['workers']}, delay={test_result['config']['delay_ms']}ms",
                '总请求': total,
                'HTTP 200': http_200_count,
                'HTTP 401': http_401_count,
                '401占比': f"{http_401_rate:.1f}%",
                '其他错误': error_counts,
                '平均响应时间': f"{avg_response_time:.2f}s",
                '总耗时': f"{test_result['elapsed']:.1f}s"
            })
        
        # 返回原始分析数据（不依赖pandas）
        return analysis


def print_analysis_table(analysis: list):
    """打印分析结果表格"""
    print(f"\n{'='*100}")
    print("分析结果汇总")
    print(f"{'='*100}\n")
    
    print(f"{'测试场景':<25} {'配置':<30} {'200':<5} {'401':<5} {'401占比':<8} {'平均响应':<12} {'总耗时'}")
    print("-" * 100)
    
    for row in analysis:
        print(f"{row['测试场景']:<25} {row['配置']:<30} {row['HTTP 200']:<5} {row['HTTP 401']:<5} "
              f"{row['401占比']:<8} {row['平均响应时间']:<12} {row['总耗时']}")


def analyze_diagnosis(analysis: list) -> str:
    """
    根据分析结果给出诊断意见
    
    Returns:
        诊断结论字符串
    """
    # 提取401率
    rates = {row['测试场景']: float(row['401占比'].rstrip('%')) for row in analysis}
    
    single_thread_no_delay_rate = rates.get('单线程无延迟', 0)
    multi_thread_10_no_delay_rate = rates.get('多线程10工作线程0ms延迟', 0)
    multi_thread_20_no_delay_rate = rates.get('多线程20工作线程0ms延迟', 0)
    single_thread_500ms_rate = rates.get('单线程500ms延迟', 0)
    single_thread_1s_rate = rates.get('单线程1s延迟', 0)
    
    diagnosis = []
    diagnosis.append("诊断分析：\n")
    
    # 判断IP封禁 vs 并发限制
    if single_thread_no_delay_rate > 50:
        diagnosis.append("❌ 【结论】：极可能是 IP 封禁（或被加入黑名单）")
        diagnosis.append(f"   - 单线程无延迟的401率高达 {single_thread_no_delay_rate:.1f}%")
        diagnosis.append("   - 即使单个请求也被拒绝，说明问题在服务器端已识别的IP")
        diagnosis.append("\n  【对策】：")
        diagnosis.append("   1. 更换客户端IP（VPN/代理）")
        diagnosis.append("   2. 等待足够长时间（可能需要数小时到数天）")
        diagnosis.append("   3. 尝试使用不同的User-Agent和Headers")
        diagnosis.append("   4. 考虑使用住宅代理池或SOCKS代理")
    
    elif (multi_thread_20_no_delay_rate - single_thread_no_delay_rate) > 20:
        diagnosis.append("✓ 【结论】：主要是 并发控制触发（目标网站有速率限制）")
        diagnosis.append(f"   - 单线程401率: {single_thread_no_delay_rate:.1f}%")
        diagnosis.append(f"   - 20线程401率: {multi_thread_20_no_delay_rate:.1f}%")
        diagnosis.append(f"   - 差异: {multi_thread_20_no_delay_rate - single_thread_no_delay_rate:.1f}%")
        diagnosis.append("\n  【对策】：")
        diagnosis.append("   1. ⭐ 降低并发数（改为2-5个工作线程）")
        diagnosis.append("   2. 在请求间加延迟（500ms-1s）")
        diagnosis.append("   3. 实现指数退避重试机制（遇到401时等待更长时间后重试）")
        diagnosis.append("   4. 添加随机延迟避免规律性请求被识别")
    
    elif (single_thread_1s_rate - single_thread_no_delay_rate) < -10:
        diagnosis.append("✓ 【结论】：目标网站实施了 速率限制")
        diagnosis.append(f"   - 无延迟401率: {single_thread_no_delay_rate:.1f}%")
        diagnosis.append(f"   - 1s延迟401率: {single_thread_1s_rate:.1f}%")
        diagnosis.append(f"   - 延迟改善: {single_thread_no_delay_rate - single_thread_1s_rate:.1f}%")
        diagnosis.append("\n  【对策】：")
        diagnosis.append("   1. 在fetch.py中添加 --min-request-interval 参数（默认500ms）")
        diagnosis.append("   2. 实现自适应延迟（遇到401时自动增加延迟）")
        diagnosis.append("   3. 降低默认workers数（从10改为3-5）")
    
    else:
        diagnosis.append("⚠️ 【结论】：情况复杂或采样量不足")
        diagnosis.append(f"   - 单线程401率: {single_thread_no_delay_rate:.1f}%")
        diagnosis.append(f"   - 多线程401率: {multi_thread_10_no_delay_rate:.1f}%")
        diagnosis.append("   - 无明显规律")
        diagnosis.append("\n  【建议】：")
        diagnosis.append("   1. 增加采样量（sample_size > 100）")
        diagnosis.append("   2. 检查目标URL是否有特殊模式（某些URL总是被拒绝）")
        diagnosis.append("   3. 查看具体返回的HTTP Headers（Location/Retry-After）")
    
    return "\n".join(diagnosis)


def save_detailed_results(all_results: list, output_path: str):
    """保存详细结果到JSON文件"""
    output_data = {
        'timestamp': datetime.now().isoformat(),
        'tests': []
    }
    
    for test in all_results:
        test_data = {
            'name': test['name'],
            'config': test['config'],
            'elapsed': test['elapsed'],
            'results': []
        }
        
        for r in test['results']:
            test_data['results'].append({
                'url': r['url'],
                'status': r['status'],
                'error': r['error'],
                'response_time': round(r['response_time'], 3)
            })
        
        output_data['tests'].append(test_data)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ 详细结果已保存到: {output_path}")


if __name__ == '__main__':
    import sys
    
    # 获取Excel文件路径
    excel_file = sys.argv[1] if len(sys.argv) > 1 else '部分.工作簿1.xlsx'
    
    if not Path(excel_file).exists():
        print(f"✗ 文件不存在: {excel_file}")
        print("\n用法: python diagnose_blocking.py <excel_file> [sample_size]")
        print("示例: python diagnose_blocking.py 部分.工作簿1.xlsx 30")
        sys.exit(1)
    
    # 采样数量
    sample_size = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    # 运行诊断
    diagnoser = RequestDiagnoser(excel_file, sample_size=sample_size)
    all_results = diagnoser.run_all_tests()
    
    # 分析结果
    analysis = RequestDiagnoser.analyze_results(all_results)
    print_analysis_table(analysis)
    
    # 给出诊断建议
    diagnosis_text = analyze_diagnosis(analysis)
    print(f"\n{diagnosis_text}")
    
    # 保存详细结果
    output_path = 'diagnose_blocking_results.json'
    save_detailed_results(all_results, output_path)
    
    print(f"\n{'='*100}")
    print("✓ 诊断完成！")
