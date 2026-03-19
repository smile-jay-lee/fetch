"""
诊断脚本：review 端点专用测试
测试/review?id=xxx端点（不是newGbInfo）
"""

import sys
import time
import re
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook


def get_urls_from_n_hint(excel_path, count=20):
    """从N列获取URL（这些是有采标信息的）"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    urls = []
    
    for row in range(2, ws.max_row + 1):
        # 检查N列是否有'采'字（N-hint标记）
        n_val = ws['N' + str(row)].value
        if n_val and isinstance(n_val, str) and '采' in n_val:
            # 从O列获取newGbInfo URL
            o_val = ws['O' + str(row)].value
            if o_val and isinstance(o_val, str):
                # 提取hcno参数
                match = re.search(r'hcno=([A-F0-9]+)', o_val)
                if match:
                    hcno = match.group(1)
                    # 构造review URL (假设ID就是hcno)
                    review_url = f"https://std.samr.gov.cn/gb/search/gbDetailed?id={hcno}&review=true"
                    urls.append((row, review_url, o_val))
                    if len(urls) >= count:
                        break
    
    return urls


def fetch_url(url, timeout=8):
    """发起单个请求"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Referer': 'https://std.samr.gov.cn/'
        }
        req = Request(url, headers=headers)
        resp = urlopen(req, timeout=timeout)
        return {'status': resp.status, 'error': None}
    except HTTPError as e:
        return {'status': e.code, 'error': f'HTTP{e.code}'}
    except Exception as e:
        return {'status': None, 'error': type(e).__name__}


if __name__ == '__main__':
    excel = sys.argv[1] if len(sys.argv) > 1 else '部分.工作簿1.xlsx'
    sample = int(sys.argv[2]) if len(sys.argv) > 2 else 20
    
    print("=" * 70)
    print("诊断脚本：review端点测试")
    print("=" * 70)
    
    urls = get_urls_from_n_hint(excel, count=sample)
    if not urls:
        print("✗ 没有找到N列有'采'字的行")
        sys.exit(1)
    
    print(f"\n✓ 加载了 {len(urls)} 个review端点URL（从N列采标行）\n")
    
    # 测试1：单线程
    print("【测试1】单线程无延迟...")
    single_results = []
    for i, (row_num, url, _) in enumerate(urls, 1):
        result = fetch_url(url)
        single_results.append(result)
        status = f"{result['status']}" if result['status'] else result['error']
        print(f"  {i}. 行{row_num}: {status}", end='')
        if result['status'] == 401:
            print(" ✗")
        elif result['status'] == 200:
            print(" ✓")
        else:
            print()
    
    time.sleep(1)
    
    # 测试2：多线程
    print(f"\n【测试2】多线程(10个工作线程)...")
    multi_results = [None] * len(urls)
    
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_url, url): i 
                  for i, (row_num, url, _) in enumerate(urls)}
        for future in as_completed(futures):
            i = futures[future]
            multi_results[i] = future.result()
    
    for i, result in enumerate(multi_results, 1):
        row_num = urls[i-1][0]
        status = f"{result['status']}" if result['status'] else result['error']
        print(f"  {i}. 行{row_num}: {status}", end='')
        if result['status'] == 401:
            print(" ✗")
        elif result['status'] == 200:
            print(" ✓")
        else:
            print()
    
    # 分析
    print("\n" + "=" * 70)
    print("分析结果")
    print("=" * 70)
    
    single_401 = sum(1 for r in single_results if r['status'] == 401)
    multi_401 = sum(1 for r in multi_results if r['status'] == 401)
    single_200 = sum(1 for r in single_results if r['status'] == 200)
    multi_200 = sum(1 for r in multi_results if r['status'] == 200)
    single_timeout = sum(1 for r in single_results if r['error'])
    multi_timeout = sum(1 for r in multi_results if r['error'])
    
    single_401_rate = (single_401 / len(urls) * 100) if urls else 0
    multi_401_rate = (multi_401 / len(urls) * 100) if urls else 0
    diff = multi_401_rate - single_401_rate
    
    print(f"\n单线程：200={single_200}个, 401={single_401}个, 其他={single_timeout}个")
    print(f"         401率: {single_401_rate:.1f}%")
    print(f"\n多线程：200={multi_200}个, 401={multi_401}个, 其他={multi_timeout}个")
    print(f"         401率: {multi_401_rate:.1f}%")
    print(f"\n401率差异：{diff:+.1f}%")
    
    print("\n【结论】：")
    if single_401_rate > 60:
        print("❌ review端点被广泛限制/封禁")
        print("   → review端点与newGbInfo端点情况不同")
        print("   → 需要使用代理或等待")
    elif multi_401_rate - single_401_rate > 20:
        print("⚠️  review端点对并发敏感")
        print("   → 需要降低workers或加延迟")
    elif single_401_rate > 30:
        print("⚠️  review端点有中等限制")
        print("   → 需要重试+延迟策略")
    else:
        print("✓ review端点正常")
        print("   → 之前的401是网络抖动")
    
    print("\n关键数据点：")
    for i, (row_num, url, newgb_url) in enumerate(urls[:5]):
        result = single_results[i]
        print(f"  行{row_num}: {result['status']} - {url[:60]}...")
    
    print("\n" + "=" * 70)
