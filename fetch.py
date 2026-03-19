import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib
import html
import re
import sys
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
from requests.adapters import HTTPAdapter
from requests.exceptions import HTTPError as RequestsHTTPError
from requests.exceptions import RequestException
from urllib3.util.retry import Retry


_HTTP_SESSION: requests.Session | None = None


def get_http_session() -> requests.Session:
	global _HTTP_SESSION
	if _HTTP_SESSION is not None:
		return _HTTP_SESSION

	retry = Retry(
		total=3,
		connect=3,
		read=3,
		status=3,
		backoff_factor=0.5,
		status_forcelist=(429, 500, 502, 503, 504),
		allowed_methods=frozenset({"HEAD", "GET", "OPTIONS"}),
		raise_on_status=False,
		respect_retry_after_header=True,
	)
	adapter = HTTPAdapter(max_retries=retry, pool_connections=100, pool_maxsize=100)

	session = requests.Session()
	session.mount("http://", adapter)
	session.mount("https://", adapter)
	_HTTP_SESSION = session
	return session


def _format_http_error(exc: RequestsHTTPError) -> str:
	resp = exc.response
	if resp is None:
		return f"{exc.__class__.__name__}: HTTP unknown body=''"
	body = (resp.text or "").replace("\n", " ").replace("\r", " ")[:100]
	return f"{exc.__class__.__name__}: HTTP {resp.status_code} body='{body}'"


def _format_network_error(exc: RequestException) -> str:
	return f"{exc.__class__.__name__}: {exc}"


class FkButtonParser(HTMLParser):
	def __init__(self) -> None:
		super().__init__()
		self.data_values: list[str] = []

	def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
		if tag.lower() != "button":
			return

		attrs_dict = {k: (v or "") for k, v in attrs}
		class_name = attrs_dict.get("class", "")
		if "fk_btn" not in class_name.split():
			return

		value = attrs_dict.get("data-value", "").strip()
		if value:
			self.data_values.append(value)


class TextExtractor(HTMLParser):
	def __init__(self) -> None:
		super().__init__()
		self.parts: list[str] = []
		self._skip_level = 0

	def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
		if tag.lower() in {"script", "style"}:
			self._skip_level += 1

	def handle_endtag(self, tag: str) -> None:
		if tag.lower() in {"script", "style"} and self._skip_level > 0:
			self._skip_level -= 1

	def handle_data(self, data: str) -> None:
		if self._skip_level == 0 and data:
			self.parts.append(data)


def find_excel_file(base_dir: Path, explicit_path: str | None) -> Path:
	if explicit_path:
		p = Path(explicit_path)
		if not p.is_absolute():
			p = base_dir / p
		if not p.exists():
			raise FileNotFoundError(f"Excel file not found: {p}")
		return p

	candidates = [
		p
		for p in base_dir.glob("*.xlsx")
		if p.is_file() and not p.name.startswith("~$")
	]
	if not candidates:
		raise FileNotFoundError("No .xlsx file found in current directory")

	candidates.sort()
	if len(candidates) > 1:
		names = ", ".join(p.name for p in candidates)
		raise RuntimeError(
			"Multiple .xlsx files found. Please specify one with --excel. "
			f"Candidates: {names}"
		)

	return candidates[0]


def save_workbook_with_fallback(wb: object, desired_path: Path) -> Path:
	try:
		wb.save(desired_path)
		return desired_path
	except PermissionError:
		ts = datetime.now().strftime("%Y%m%d_%H%M%S")
		fallback = desired_path.with_name(f"{desired_path.stem}.autosave_{ts}{desired_path.suffix}")
		wb.save(fallback)
		print(
			f"[WARN] Cannot write '{desired_path}' (likely open in Excel). "
			f"Saved to '{fallback}' instead."
		)
		return fallback


def normalize_url(value: object) -> str | None:
	if value is None:
		return None
	text = str(value).strip()
	if not text:
		return None

	# Ignore common header-like placeholders.
	if text.lower() in {"url", "article_url", "link"}:
		return None

	if not text.lower().startswith(("http://", "https://")):
		text = "http://" + text

	parsed = urlparse(text)
	if not parsed.netloc:
		return None

	host = parsed.netloc.split("@")[-1].split(":")[0]
	if "." not in host and host.lower() != "localhost":
		return None

	return text


def build_filename(url: str, row: int) -> str:
	parsed = urlparse(url)
	raw = f"{parsed.netloc}{parsed.path}".strip("/")
	slug = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("_")
	if not slug:
		slug = "page"
	slug = slug[:80]
	digest = hashlib.md5(url.encode("utf-8"), usedforsecurity=False).hexdigest()[:8]
	return f"row{row:04d}_{slug}_{digest}.html"


def build_request_headers(url: str) -> dict[str, str]:
	headers = {
		"User-Agent": (
			"Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
			"AppleWebKit/537.36 (KHTML, like Gecko) "
			"Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
		),
		"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
		"Accept-Encoding": "gzip, deflate, br",
		"Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
		"Cache-Control": "no-cache",
		"Pragma": "no-cache",
		"Sec-Fetch-Dest": "document",
		"Sec-Fetch-Mode": "navigate",
		"Sec-Fetch-Site": "none",
		"Sec-Fetch-User": "?1",
		"Sec-Ch-Ua": '"Chromium";v="124", "Google Chrome";v="124", ";Not A Brand";v="99"',
		"Sec-Ch-Ua-Mobile": "?0",
		"Sec-Ch-Ua-Platform": '"Windows"',
		"Upgrade-Insecure-Requests": "1",
	}

	host = urlparse(url).netloc.lower()
	if "std.samr.gov.cn" in host:
		headers["Referer"] = "https://std.samr.gov.cn/"
	elif "openstd.samr.gov.cn" in host:
		headers["Referer"] = "https://openstd.samr.gov.cn/"

	return headers


def remove_review_true(url: str) -> str:
	parsed = urlparse(url)
	query_pairs = parse_qsl(parsed.query, keep_blank_values=True)
	filtered = [
		(k, v)
		for k, v in query_pairs
		if not (k.lower() == "review" and v.lower() == "true")
	]
	if len(filtered) == len(query_pairs):
		return url
	return urlunparse(parsed._replace(query=urlencode(filtered)))


def fetch_url_content(url: str, timeout: int) -> str:
	resp = get_http_session().get(url, headers=build_request_headers(url), timeout=timeout)
	resp.raise_for_status()
	if not resp.encoding:
		resp.encoding = "utf-8"
	return resp.text


def fetch_url_content_resilient(url: str, timeout: int) -> tuple[str, str]:
	try:
		return fetch_url_content(url, timeout), url
	except RequestsHTTPError as exc:
		fallback_url = remove_review_true(url)
		status_code = exc.response.status_code if exc.response is not None else 0
		if status_code >= 500 and fallback_url != url:
			return fetch_url_content(fallback_url, timeout), fallback_url
		raise


def html_to_text(raw_html: str) -> str:
	parser = TextExtractor()
	parser.feed(raw_html)
	text = " ".join(parser.parts)
	text = html.unescape(text)
	text = text.replace("\u3000", " ")
	text = re.sub(r"\s+", " ", text)
	return text.strip()


def extract_section_text_by_title(raw_html: str, section_title: str) -> str:
	# Match the section block from the title to the next para-title/footer/body end.
	pattern = re.compile(
		rf'<h2[^>]*class="title-text"[^>]*>\s*{re.escape(section_title)}\s*</h2>([\s\S]*?)(?=<div\s+class="para-title"|<div\s+class="footer-body"|</body>)',
		flags=re.IGNORECASE,
	)
	match = pattern.search(raw_html)
	if not match:
		return ""

	section_html = match.group(1)
	section_text = html_to_text(section_html)
	# Remove repeated section title if present in text extraction.
	section_text = section_text.replace(section_title, "").strip()
	return section_text


def extract_caibiao_module_raw(raw_html: str) -> str:
	section_text = extract_section_text_by_title(raw_html, "采标情况")
	if section_text:
		return section_text[:1000].strip()

	text = html_to_text(raw_html)
	if "采标情况" not in text:
		return ""

	# Fallback: capture a bounded snippet after the section title.
	idx = text.find("采标情况")
	snippet = text[idx: idx + 600]
	snippet = re.split(r"(起草单位|起草人|归口单位|发布日期|实施日期|ICS|中国标准分类号)", snippet)[0]
	return snippet.strip()


def extract_caibiao_details(module_raw: str) -> tuple[str, str]:
	if not module_raw:
		return "", ""
	
	degree = ""
	target_obj = ""
	
	# Match degree keyword (Chinese or English adoption codes)
	match_degree = re.search(r"(等同采用|修改采用|非等效采用|IDT|MOD|NEQ)", module_raw)
	if match_degree:
		degree = match_degree.group(1).upper() if match_degree.group(1) in ("IDT", "MOD", "NEQ") else match_degree.group(1)
		# Extract target object after degree keyword (二次切片)
		tail = module_raw[match_degree.end():].split("。")[0].strip()
		# Remove common prefixes
		target_obj = re.sub(r"^(了|的)?(IEC|ISO|EN|ITU)?(国际)?标准[：:]?\s*", "", tail).strip()
	
	return degree, target_obj


def extract_caibiao_info(raw_html: str) -> tuple[str, str, str]:
	module_raw = extract_caibiao_module_raw(raw_html)
	degree, target_obj = extract_caibiao_details(module_raw)
	return module_raw, degree, target_obj


def resolve_caibiao_column_index(ws: object, column_spec: str) -> int:
	if column_spec.upper() == "AUTO":
		return ws.max_column + 1

	from openpyxl.utils import column_index_from_string

	return column_index_from_string(column_spec.upper())


def resolve_output_column_index(ws: object, column_spec: str, header_name: str | None) -> int:
	if column_spec.upper() != "AUTO":
		return resolve_caibiao_column_index(ws, column_spec)

	if header_name:
		for col in range(1, ws.max_column + 1):
			v = ws.cell(row=1, column=col).value
			if v is not None and str(v).strip() == header_name:
				return col

	return ws.max_column + 1


def n_column_hint_has_caibiao(value: object) -> bool:
	if value is None:
		return False
	text = str(value).strip()
	if not text:
		return False
	return "采" in text


def extract_feedback_ids(html: str) -> list[str]:
	parser = FkButtonParser()
	parser.feed(html)

	# Keep order while de-duplicating.
	ordered_unique: list[str] = []
	seen: set[str] = set()
	for value in parser.data_values:
		if value not in seen:
			ordered_unique.append(value)
			seen.add(value)
	return ordered_unique


def build_review_url(base_url: str, hcno: str) -> str:
	parsed = urlparse(base_url)
	origin = f"{parsed.scheme}://{parsed.netloc}"
	return urljoin(origin, f"/bzgk/gb/review?hcno={hcno}")


def build_review_filename(row: int, hcno: str) -> str:
	safe_hcno = re.sub(r"[^A-Za-z0-9._-]+", "_", hcno).strip("_")[:80] or "unknown"
	return f"row{row:04d}_review_{safe_hcno}.html"


def extract_hcnos_for_review(url: str, html_content: str) -> list[str]:
	ids = extract_feedback_ids(html_content)
	if ids:
		return ids

	parsed = urlparse(url)
	query = dict(parse_qsl(parsed.query, keep_blank_values=True))
	hcno = query.get("hcno", "").strip()
	if hcno:
		return [hcno]
	return []


def process_row_task(
	row: int,
	url: str,
	output_dir: Path,
	review_dir: Path,
	timeout: int,
	fetch_review: bool,
	extract_caibiao: bool,
) -> dict[str, object]:
	result: dict[str, object] = {
		"row": row,
		"url": url,
		"ok": False,
		"error": "",
		"final_url": url,
		"saved_name": "",
		"predicted_has_caibiao": False,
		"caibiao_degree": "",
		"caibiao_module_raw": "",
		"caibiao_object": "",
		"caibiao_from_review": False,
		"review_found": 0,
		"review_success": 0,
		"review_failed": 0,
		"review_fail_msgs": [],
	}

	filename = build_filename(url, row)
	target = output_dir / filename

	try:
		content, final_url = fetch_url_content_resilient(url, timeout=timeout)
		target.write_text(content, encoding="utf-8")

		result["ok"] = True
		result["final_url"] = final_url
		result["saved_name"] = target.name

		hcnos = extract_feedback_ids(content)
		result["review_found"] = len(hcnos)

		review_cache: dict[str, tuple[str, str]] = {}
		review_saved_hcnos: set[str] = set()

		def fetch_review_by_hcno(hcno: str, resilient: bool) -> tuple[str, str, str]:
			if hcno in review_cache:
				rc, rf = review_cache[hcno]
				ru = build_review_url(url, hcno)
				return rc, rf, ru

			review_url = build_review_url(url, hcno)
			if resilient:
				rc, rf = fetch_url_content_resilient(review_url, timeout=timeout)
			else:
				rc = fetch_url_content(review_url, timeout=timeout)
				rf = review_url
			review_cache[hcno] = (rc, rf)
			return rc, rf, review_url

		if extract_caibiao:
			best_module_raw, best_caibiao_degree, best_caibiao_object = extract_caibiao_info(content)
			predicted_has_caibiao = bool(best_caibiao_degree)
			caibiao_from_review = False

			if not predicted_has_caibiao:
				fallback_hcnos = extract_hcnos_for_review(url, content)
				for hcno in fallback_hcnos:
					try:
						rc, _, review_url = fetch_review_by_hcno(hcno, resilient=True)
						r_module_raw, r_degree, r_object = extract_caibiao_info(rc)
						if r_degree:
							predicted_has_caibiao = True
							if r_module_raw:
								best_module_raw = r_module_raw
							best_caibiao_degree = r_degree
							best_caibiao_object = r_object
							caibiao_from_review = True
							if fetch_review:
								review_target = review_dir / build_review_filename(row, hcno)
								review_target.write_text(rc, encoding="utf-8")
								review_saved_hcnos.add(hcno)
							break
						if (not best_module_raw) and r_module_raw:
							best_module_raw = r_module_raw
					except RequestsHTTPError as review_http_exc:
						print(
							f"[HTTP ERROR][REVIEW-FALLBACK] row={row} hcno={hcno} "
							f"url={review_url} {_format_http_error(review_http_exc)}"
						)
						continue
					except RequestException as review_net_exc:
						print(
							f"[NET ERROR][REVIEW-FALLBACK] row={row} hcno={hcno} "
							f"url={review_url} {_format_network_error(review_net_exc)}"
						)
						continue
					except Exception as review_exc:
						print(
							f"[ERROR][REVIEW-FALLBACK] row={row} hcno={hcno} "
							f"url={review_url} {review_exc.__class__.__name__}: {review_exc}"
						)
						continue

			result["predicted_has_caibiao"] = predicted_has_caibiao
			result["caibiao_degree"] = best_caibiao_degree
			result["caibiao_module_raw"] = best_module_raw
			result["caibiao_object"] = best_caibiao_object
			result["caibiao_from_review"] = caibiao_from_review

		if fetch_review:
			review_success = 0
			review_failed = 0
			review_fail_msgs: list[str] = []
			for hcno in hcnos:
				if hcno in review_saved_hcnos:
					review_success += 1
					continue
				review_url = build_review_url(url, hcno)
				review_target = review_dir / build_review_filename(row, hcno)
				try:
					if hcno in review_cache:
						rc, _ = review_cache[hcno]
					else:
						rc = fetch_url_content(review_url, timeout=timeout)
					review_target.write_text(rc, encoding="utf-8")
					review_success += 1
				except RequestsHTTPError as review_http_exc:
					review_failed += 1
					review_fail_msgs.append(
						f"row={row} url={review_url} error={_format_http_error(review_http_exc)}"
					)
				except RequestException as review_net_exc:
					review_failed += 1
					review_fail_msgs.append(
						f"row={row} url={review_url} error={_format_network_error(review_net_exc)}"
					)
				except Exception as review_exc:
					review_failed += 1
					review_fail_msgs.append(
						f"row={row} url={review_url} error={review_exc.__class__.__name__}: {review_exc}"
					)

			result["review_success"] = review_success
			result["review_failed"] = review_failed
			result["review_fail_msgs"] = review_fail_msgs

	except RequestsHTTPError as exc:
		msg = _format_http_error(exc)
		print(f"[HTTP ERROR] row={row} url={url} {msg}")
		result["error"] = msg
	except RequestException as exc:
		msg = _format_network_error(exc)
		print(f"[NET ERROR] row={row} url={url} {msg}")
		result["error"] = msg
	except Exception as exc:
		msg = f"{exc.__class__.__name__}: {exc}"
		print(f"[ERROR] row={row} url={url} {msg}")
		result["error"] = msg

	return result


def run(
	excel_path: Path,
	output_dir: Path,
	sheet_name: str | None,
	timeout: int,
	workers: int,
	max_rows: int | None,
	fetch_review: bool,
	extract_caibiao: bool,
	caibiao_column: str,
	caibiao_header: str,
	excel_output: Path | None,
	enable_n_hint_validation: bool,
	n_hint_column: str,
	n_hint_validation_column: str,
	n_hint_validation_header: str,
) -> int:
	try:
		from openpyxl import load_workbook
	except ImportError as exc:
		raise RuntimeError(
			"Missing dependency: openpyxl. Install with: pip install openpyxl"
		) from exc

	wb = load_workbook(excel_path)
	ws = wb[sheet_name] if sheet_name else wb.active

	caibiao_col_idx: int | None = None
	caibiao_obj_col_idx: int | None = None
	caibiao_raw_col_idx: int | None = None
	n_hint_col_idx: int | None = None
	n_hint_validation_col_idx: int | None = None
	if extract_caibiao:
		caibiao_col_idx = resolve_output_column_index(ws, caibiao_column, caibiao_header)
		if caibiao_header and not ws.cell(row=1, column=caibiao_col_idx).value:
			ws.cell(row=1, column=caibiao_col_idx, value=caibiao_header)

		caibiao_obj_col_idx = resolve_output_column_index(ws, "AUTO", "采用标准对象")
		if not ws.cell(row=1, column=caibiao_obj_col_idx).value:
			ws.cell(row=1, column=caibiao_obj_col_idx, value="采用标准对象")

		caibiao_raw_col_idx = resolve_output_column_index(ws, "AUTO", "采标模块原文(保底)")
		if not ws.cell(row=1, column=caibiao_raw_col_idx).value:
			ws.cell(row=1, column=caibiao_raw_col_idx, value="采标模块原文(保底)")

		n_hint_col_idx = resolve_caibiao_column_index(ws, n_hint_column)
		n_hint_validation_col_idx = resolve_output_column_index(
			ws,
			"AUTO",
			n_hint_validation_header,
		)
		if n_hint_validation_header and not ws.cell(row=1, column=n_hint_validation_col_idx).value:
			ws.cell(row=1, column=n_hint_validation_col_idx, value=n_hint_validation_header)
	elif enable_n_hint_validation:
		n_hint_col_idx = resolve_caibiao_column_index(ws, n_hint_column)
		if n_hint_validation_column.upper() != "NONE":
			n_hint_validation_col_idx = resolve_caibiao_column_index(ws, n_hint_validation_column)
			if n_hint_validation_header and not ws.cell(row=1, column=n_hint_validation_col_idx).value:
				ws.cell(row=1, column=n_hint_validation_col_idx, value=n_hint_validation_header)

	output_dir.mkdir(parents=True, exist_ok=True)
	review_dir = output_dir / "review"
	if fetch_review:
		review_dir.mkdir(parents=True, exist_ok=True)

	success = 0
	failed = 0
	processed = 0
	review_found = 0
	review_success = 0
	review_failed = 0
	caibiao_module_found = 0
	caibiao_text_found = 0
	caibiao_from_review_found = 0
	n_hint_positive = 0
	n_hint_positive_and_predicted = 0
	n_hint_positive_but_not_predicted = 0
	n_hint_negative_but_predicted = 0

	if workers < 1:
		raise ValueError("workers must be >= 1")

	row_jobs: list[tuple[int, str]] = []
	for row in range(1, ws.max_row + 1):
		if max_rows is not None and len(row_jobs) >= max_rows:
			break
		raw = ws[f"O{row}"].value
		url = normalize_url(raw)
		if not url:
			continue
		row_jobs.append((row, url))

	processed = len(row_jobs)

	with ThreadPoolExecutor(max_workers=workers) as executor:
		future_to_job = {
			executor.submit(
				process_row_task,
				row,
				url,
				output_dir,
				review_dir,
				timeout,
				fetch_review,
				extract_caibiao,
			): (row, url)
			for row, url in row_jobs
		}

		for future in as_completed(future_to_job):
			row, url = future_to_job[future]
			try:
				res = future.result()
			except Exception as exc:
				res = {
					"row": row,
					"url": url,
					"ok": False,
					"error": f"{exc.__class__.__name__}: {exc}",
					"predicted_has_caibiao": False,
					"caibiao_degree": "",
					"caibiao_module_raw": "",
					"caibiao_object": "",
					"caibiao_from_review": False,
					"review_found": 0,
					"review_success": 0,
					"review_failed": 0,
					"review_fail_msgs": [],
				}

			ok = bool(res.get("ok", False))
			if ok:
				success += 1
				saved_name = str(res.get("saved_name", ""))
				final_url = str(res.get("final_url", url))
				print(f"[OK] row={row} saved={saved_name} url={final_url}")
			else:
				failed += 1
				err = str(res.get("error", "unknown error"))
				print(f"[FAIL] row={row} url={url} error={err}")

			if fetch_review:
				review_found += int(res.get("review_found", 0))
				review_success += int(res.get("review_success", 0))
				review_failed += int(res.get("review_failed", 0))
				for msg in res.get("review_fail_msgs", []):
					print(f"[FAIL][REVIEW] {msg}")

			if extract_caibiao and caibiao_col_idx is not None:
				caibiao_degree = str(res.get("caibiao_degree", "")).strip()
				caibiao_module_raw = str(res.get("caibiao_module_raw", "")).strip()
				caibiao_object = str(res.get("caibiao_object", "")).strip()
				if bool(res.get("caibiao_from_review", False)):
					caibiao_from_review_found += 1

				if caibiao_module_raw:
					caibiao_module_found += 1
				if caibiao_degree:
					caibiao_text_found += 1

				if caibiao_raw_col_idx is not None:
					ws.cell(row=row, column=caibiao_raw_col_idx, value=caibiao_module_raw)

				# Write degree keyword
				if caibiao_degree:
					ws.cell(row=row, column=caibiao_col_idx, value=caibiao_degree)
				else:
					ws.cell(row=row, column=caibiao_col_idx, value="")

				# Write standard object
				if caibiao_obj_col_idx is not None:
					if caibiao_object:
						ws.cell(row=row, column=caibiao_obj_col_idx, value=caibiao_object)
					else:
						ws.cell(row=row, column=caibiao_obj_col_idx, value="")

			if n_hint_col_idx is not None:
				n_hint_value = ws.cell(row=row, column=n_hint_col_idx).value
				expected_by_hint = n_column_hint_has_caibiao(n_hint_value)
				predicted_has_caibiao = bool(str(res.get("caibiao_degree", "")).strip())

				if expected_by_hint:
					n_hint_positive += 1
					if predicted_has_caibiao:
						n_hint_positive_and_predicted += 1
					else:
						n_hint_positive_but_not_predicted += 1
				elif predicted_has_caibiao:
					n_hint_negative_but_predicted += 1

				if n_hint_validation_col_idx is not None:
					if predicted_has_caibiao:
						status = "N列命中且已提取" if expected_by_hint else "N列未标采但提取为采标"
					else:
						# FIXED: Output meaningful status instead of empty string
						status = "N列疑似采标但未提取" if expected_by_hint else "N列未标采且未提取"
					ws.cell(row=row, column=n_hint_validation_col_idx, value=status)

	if extract_caibiao:
		out_excel = excel_output or excel_path
		out_excel = save_workbook_with_fallback(wb, out_excel)

	print("\nDone")
	print(f"Excel: {excel_path}")
	print(f"Sheet: {ws.title}")
	print(f"Processed URLs: {processed}")
	print(f"Success: {success}")
	print(f"Failed: {failed}")
	if fetch_review:
		print(f"Review IDs found: {review_found}")
		print(f"Review success: {review_success}")
		print(f"Review failed: {review_failed}")
		print(f"Review output folder: {review_dir}")
	if extract_caibiao:
		print(f"Caibiao module found: {caibiao_module_found}")
		print(f"Caibiao text extracted: {caibiao_text_found}")
		print(f"Caibiao hits from review fallback: {caibiao_from_review_found}")
	if n_hint_col_idx is not None:
		print(f"N-hint positive (N列含'采'): {n_hint_positive}")
		print(f"N-hint matched by extraction: {n_hint_positive_and_predicted}")
		print(f"N-hint possible misses: {n_hint_positive_but_not_predicted}")
		print(f"N-hint negative but extracted: {n_hint_negative_but_predicted}")
	if extract_caibiao:
		print(f"Excel updated: {out_excel}")
	print(f"Output folder: {output_dir}")
	if failed == 0 and (not fetch_review or review_failed == 0):
		return 0
	return 1


def main() -> int:
	parser = argparse.ArgumentParser(
		description="Read URLs from column O in Excel and save webpage content to content/ folder"
	)
	parser.add_argument("--excel", help="Excel file path (default: first .xlsx in current folder)")
	parser.add_argument("--sheet", help="Sheet name (default: active sheet)")
	parser.add_argument("--timeout", type=int, default=15, help="Request timeout in seconds")
	parser.add_argument("--workers", type=int, default=10, help="Concurrent worker count (default: 10)")
	parser.add_argument("--max", type=int, default=None, help="Only process first N non-empty URLs")
	parser.add_argument(
		"--no-review",
		action="store_true",
		help="Only fetch URLs in column O, do not fetch /bzgk/gb/review pages",
	)
	parser.add_argument(
		"--no-caibiao",
		action="store_true",
		help="Skip checking/extracting \"采标情况\" information",
	)
	parser.add_argument(
		"--caibiao-column",
		default="AUTO",
		help="Target Excel column for extracted 采标情况 text (e.g. P, Q, AUTO)",
	)
	parser.add_argument(
		"--caibiao-header",
		default="采标情况提取",
		help="Header name for the output 采标情况 column",
	)
	parser.add_argument(
		"--excel-out",
		help="Output Excel path (default: overwrite source Excel)",
	)
	parser.add_argument(
		"--n-hint-validation",
		action="store_true",
		help="Enable internal validation: use column N containing '采' as a weak positive hint",
	)
	parser.add_argument(
		"--n-hint-column",
		default="N",
		help="Hint source column for weak label validation (default: N)",
	)
	parser.add_argument(
		"--n-hint-validation-column",
		default="NONE",
		help="Write per-row validation status to this column (e.g. R). Use NONE to skip writing",
	)
	parser.add_argument(
		"--n-hint-validation-header",
		default="采标校验(N列弱标签)",
		help="Header name for N-hint validation output column",
	)
	args = parser.parse_args()

	base_dir = Path(__file__).resolve().parent
	try:
		excel_path = find_excel_file(base_dir, args.excel)
		output_dir = base_dir / "content"
		excel_out = None
		if args.excel_out:
			excel_out = Path(args.excel_out)
			if not excel_out.is_absolute():
				excel_out = base_dir / excel_out
		return run(
			excel_path=excel_path,
			output_dir=output_dir,
			sheet_name=args.sheet,
			timeout=args.timeout,
			workers=args.workers,
			max_rows=args.max,
			fetch_review=not args.no_review,
			extract_caibiao=not args.no_caibiao,
			caibiao_column=args.caibiao_column,
			caibiao_header=args.caibiao_header,
			excel_output=excel_out,
			enable_n_hint_validation=args.n_hint_validation,
			n_hint_column=args.n_hint_column,
			n_hint_validation_column=args.n_hint_validation_column,
			n_hint_validation_header=args.n_hint_validation_header,
		)
	except Exception as exc:
		print(f"Error: {exc}")
		return 2


if __name__ == "__main__":
	sys.exit(main())
